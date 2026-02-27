"""
ActivityTrackerV3.py
----------------------------------------------------------
Feature Activity Checklist Tracker  —  v3

Changes from v2:
  - Export now produces a styled .xlsx (Excel) file instead of plain CSV.
  - Rows are color-coded by status:
      • Past Due   → Red fill  (#FF0000)  + Black text
      • Not Needed → Black fill (#000000) + White text
      • Complete   → Green fill (#92D050) + Black text
      • Default    → White fill           + Black text
  - Header row: Navy blue (#1F3864) fill + White bold text.
  - Phase column has its own colored background to match the app theme.
  - Dates are formatted MM/DD/YYYY; booleans rendered as Yes / No.
  - Columns are auto-sized for readability.

Run:
  streamlit run ActivityTrackerV3.py
"""

import io
import json
import os
import uuid
import streamlit as st
import pandas as pd
from datetime import date, datetime

# openpyxl styling imports
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# THEME / STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────
PHASE_COLORS = {
    "DISCOVER": "#1F6B2E",
    "DESIGN":   "#1F3864",
    "DELIVER":  "#1F6B2E",
}

PHASE_ICONS = {
    "DISCOVER": "🔍",
    "DESIGN":   "🎨",
    "DELIVER":  "🚀",
}

RESPONSIBLE_OPTIONS = [
    "",                  # blank = unassigned
    "Adam Crouch",
    "Asha Nikam",
    "John Ellison",
    "KC Chavez",
    "Kris Burch",
    "Victor Felisbino",
]

DATA_FILE = os.path.join(os.path.dirname(__file__), "activity_data_v2.json")

# ─────────────────────────────────────────────────────────────
# DEFAULT ACTIVITY DATA  (mirrors the spreadsheet)
# ─────────────────────────────────────────────────────────────
DEFAULT_ACTIVITIES = [
    # ── DISCOVER ──────────────────────────────────────────────
    {"id": 1,  "phase": "DISCOVER", "feature_activity": "Gather Existing Documentation", "purpose": "Current Process Flows",                                                                "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 2,  "phase": "DISCOVER", "feature_activity": "Gather Existing Documentation", "purpose": "System Diagrams",                                                                     "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 3,  "phase": "DISCOVER", "feature_activity": "Gather Existing Documentation", "purpose": "Architecture Review - Discovery (On the Feature Activity Checklist Spreadsheet)",      "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 4,  "phase": "DISCOVER", "feature_activity": "Gather Existing Documentation", "purpose": "Current Reporting",                                                                    "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 5,  "phase": "DISCOVER", "feature_activity": "Gather Existing Documentation", "purpose": "Past Project Information",                                                             "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 6,  "phase": "DISCOVER", "feature_activity": "Socialize Feature w/ Team",     "purpose": "Review Concept w/ Team (BA, PO, Architect, Dev and Admin)",                           "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},

    # ── DESIGN ────────────────────────────────────────────────
    {"id": 7,  "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "Updated Process Flows",                                                                "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 8,  "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "Updated System Diagrams",                                                              "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 9,  "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "Integration Data Map",                                                                 "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 10, "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "UI Mock ups (if applicable)",                                                          "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 11, "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "POC (if applicable)",                                                                  "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 12, "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "Future Reporting Impacts/Requirements (if applicable)",                                "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 13, "phase": "DESIGN",   "feature_activity": "Target State Documentation",    "purpose": "Architecture Blueprint",                                                               "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 14, "phase": "DESIGN",   "feature_activity": "Solution Review w/ Customer",   "purpose": "Review future flows and mock ups with Customer for alignment of end solution",          "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 15, "phase": "DESIGN",   "feature_activity": "Team Story Gap Session",         "purpose": "Create shell stories and \"notinachi\"",                                                "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 16, "phase": "DESIGN",   "feature_activity": "Team Story Gap Session",         "purpose": "Review Stories w/ Team",                                                               "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 17, "phase": "DESIGN",   "feature_activity": "Groom User Stories",             "purpose": "Prioritize User Stories",                                                              "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 18, "phase": "DESIGN",   "feature_activity": "Groom User Stories",             "purpose": "15% - 20% of User Stories Groomed",                                                    "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},

    # ── DELIVER ───────────────────────────────────────────────
    {"id": 19, "phase": "DELIVER",  "feature_activity": "UAT Test Plan",                          "purpose": "What do we need the business to sign off on prior to release?",                    "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 20, "phase": "DELIVER",  "feature_activity": "Training Plan",                          "purpose": "Who's responsible and to what level",                                              "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 21, "phase": "DELIVER",  "feature_activity": "Communication Plan",                     "purpose": "Ceremonies, weekly updates, demos",                                                "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 22, "phase": "DELIVER",  "feature_activity": "Video/Audio Recording of End to End Process", "purpose": "Useful reference stored in Doc Hub and linked on the Feature.",               "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 23, "phase": "DELIVER",  "feature_activity": "Project Retro",                          "purpose": "Look back on the entire project and identify what to start, stop and continue",    "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
    {"id": 24, "phase": "DELIVER",  "feature_activity": "Hypercare - Think about this",           "purpose": "Determine if/when hypercare will take place",                                      "responsible": "", "start_date": "", "due_date": "", "notes": "", "complete": False},
]


def make_default_project(name: str = "New Project") -> dict:
    """Return a fresh project dict populated with default activities."""
    acts = [a.copy() for a in DEFAULT_ACTIVITIES]
    for a in acts:
        a.setdefault("not_needed", False)
    return {
        "id":         str(uuid.uuid4()),
        "name":       name,
        "activities": acts,
    }


# ─────────────────────────────────────────────────────────────
# PERSISTENCE HELPERS
# ─────────────────────────────────────────────────────────────
def load_data() -> dict:
    """Load projects from the JSON file, or return a default structure."""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if "projects" in data:
                for p in data["projects"]:
                    for a in p["activities"]:
                        a.setdefault("not_needed", False)
                return data
        except Exception:
            pass
    return {
        "projects":   [make_default_project("My First Project")],
        "last_saved": None,
    }


def save_data(projects: list) -> str:
    """Persist all projects and return the timestamp string."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump({"projects": projects, "last_saved": ts}, f, indent=2, default=str)
    return ts


def next_id(activities: list[dict]) -> int:
    return max((a["id"] for a in activities), default=0) + 1


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT HELPERS
# ─────────────────────────────────────────────────────────────

# ── Fill presets ─────────────────────────────────────────────
_FILL_HEADER    = PatternFill("solid", fgColor="1F3864")   # Navy  – header row
_FILL_PAST_DUE  = PatternFill("solid", fgColor="FF0000")   # Red   – past due
_FILL_NOT_NEEDED = PatternFill("solid", fgColor="000000")  # Black – not needed
_FILL_COMPLETE  = PatternFill("solid", fgColor="92D050")   # Green – complete
_FILL_DEFAULT   = PatternFill("solid", fgColor="FFFFFF")   # White – default
_FILL_DISCOVER  = PatternFill("solid", fgColor="1F6B2E")   # Dark green – DISCOVER/DELIVER phase col
_FILL_DESIGN    = PatternFill("solid", fgColor="1F3864")   # Navy        – DESIGN phase col

# ── Font presets ──────────────────────────────────────────────
_FONT_HEADER      = Font(name="Segoe UI", bold=True,  color="FFFFFF", size=11)
_FONT_WHITE       = Font(name="Segoe UI", bold=False, color="FFFFFF", size=10)
_FONT_BLACK       = Font(name="Segoe UI", bold=False, color="000000", size=10)
_FONT_PHASE_WHITE = Font(name="Segoe UI", bold=True,  color="FFFFFF", size=10)

# ── Thin border ───────────────────────────────────────────────
_thin = Side(border_style="thin", color="CCCCCC")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")


def _hex(color: str) -> str:
    """Strip leading # from a hex color string."""
    return color.lstrip("#")


def _phase_fill(phase: str) -> PatternFill:
    color = _hex(PHASE_COLORS.get(phase, "#1F3864"))
    return PatternFill("solid", fgColor=color)


def _row_status(a: dict, today: date) -> str:
    """Return the status key that drives row color (priority: not_needed > past_due > complete)."""
    if a.get("not_needed", False):
        return "not_needed"
    due = _parse_date_val(a.get("due_date", ""))
    if due and due < today and not a.get("complete", False):
        return "past_due"
    if a.get("complete", False):
        return "complete"
    return "default"


def _parse_date_val(val) -> date | None:
    if not val:
        return None
    if isinstance(val, date):
        return val
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(str(val), fmt).date()
        except ValueError:
            pass
    return None


def _fmt_date_val(val) -> str:
    d = _parse_date_val(val)
    return d.strftime("%m/%d/%Y") if d else ""


def _apply_row_style(ws, row_num: int, num_cols: int, status: str, phase_col_idx: int | None = None) -> None:
    """Apply fill + font to every cell in a data row."""
    fill_map = {
        "not_needed": (_FILL_NOT_NEEDED, _FONT_WHITE),
        "past_due":   (_FILL_PAST_DUE,   _FONT_BLACK),
        "complete":   (_FILL_COMPLETE,    _FONT_BLACK),
        "default":    (_FILL_DEFAULT,     _FONT_BLACK),
    }
    fill, font = fill_map.get(status, (_FILL_DEFAULT, _FONT_BLACK))

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        # Phase column always gets phase-themed background (unless overridden by status)
        if col == phase_col_idx and status == "default":
            # leave phase col white for non-special rows so text is readable
            cell.fill = _FILL_DEFAULT
            cell.font = _FONT_BLACK
        else:
            cell.fill = fill
            cell.font = font
        cell.border = _BORDER
        cell.alignment = _ALIGN_WRAP


def _write_sheet(ws, rows: list[dict], col_names: list[str], today: date) -> None:
    """
    Write a header row + data rows with color coding to the given worksheet.
    `col_names` is the ordered list of column display names.
    `rows` is a list of dicts matching the activity structure (may include a 'project' key).
    """
    # ── Column definitions (internal_key, display_name, width) ──
    # We rebuild columns from col_names to keep things flexible
    # Always write in this canonical order:
    COLS = []
    if any("project" in r for r in rows):
        COLS.append(("project",          "Project",          22))
    COLS += [
        ("phase",            "Phase",            12),
        ("feature_activity", "Feature Activity", 30),
        ("purpose",          "Purpose",          40),
        ("responsible",      "Responsible",      18),
        ("start_date",       "Start Date",       14),
        ("due_date",         "Due Date",         14),
        ("complete",         "Complete",         10),
        ("not_needed",       "Not Needed",       12),
        ("notes",            "Notes",            45),
    ]

    num_cols = len(COLS)
    phase_col_idx = next((i + 1 for i, (k, _, _) in enumerate(COLS) if k == "phase"), None)

    # ── Header row ───────────────────────────────────────────
    for col_idx, (_, display, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.fill   = _FILL_HEADER
        cell.font   = _FONT_HEADER
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────
    for row_idx, a in enumerate(rows, start=2):
        status = _row_status(a, today)

        for col_idx, (key, _, _) in enumerate(COLS, start=1):
            raw = a.get(key, "")
            # Format dates
            if key in ("start_date", "due_date"):
                value = _fmt_date_val(raw)
            # Format booleans
            elif key in ("complete", "not_needed"):
                value = "Yes" if raw else "No"
            else:
                value = raw if raw is not None else ""
            ws.cell(row=row_idx, column=col_idx, value=value)

        _apply_row_style(ws, row_idx, num_cols, status, phase_col_idx)

    # Freeze the header row
    ws.freeze_panes = "A2"


def build_excel_bytes(activities: list[dict], sheet_name: str = "Activities") -> bytes:
    """
    Build a styled Excel workbook for a single project's activities.
    Returns the workbook as raw bytes suitable for st.download_button.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name max 31 chars

    today = date.today()
    _write_sheet(ws, activities, [], today)

    # ── Legend sheet ─────────────────────────────────────────
    _add_legend_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_bytes_all(projects: list[dict]) -> bytes:
    """
    Build a styled Excel workbook with one sheet per project + a legend sheet.
    Returns bytes for st.download_button.
    """
    wb = Workbook()
    # Remove the default empty sheet
    default_ws = wb.active
    wb.remove(default_ws)

    today = date.today()

    for proj in projects:
        # Sanitize sheet name: Excel allows max 31 chars, no special chars
        safe = proj["name"][:28].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
        ws = wb.create_sheet(title=safe)
        _write_sheet(ws, proj["activities"], [], today)

    _add_legend_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_legend_sheet(wb: Workbook) -> None:
    """Append a 'Legend' sheet explaining the color codes."""
    ws = wb.create_sheet(title="Legend")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35

    # Title
    title_cell = ws.cell(row=1, column=1, value="Color Legend")
    title_cell.font = Font(name="Segoe UI", bold=True, size=13, color="1F3864")
    ws.merge_cells("A1:B1")

    entries = [
        ("Past Due",    "FF0000", "000000", "Due date has passed and not yet complete"),
        ("Not Needed",  "000000", "FFFFFF", "Marked as not needed / N/A"),
        ("Complete",    "92D050", "000000", "Activity has been completed"),
        ("Default",     "FFFFFF", "000000", "Active — no special status"),
    ]

    for i, (label, bg, fg, desc) in enumerate(entries, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.fill   = PatternFill("solid", fgColor=bg)
        label_cell.font   = Font(name="Segoe UI", bold=True, color=fg, size=10)
        label_cell.border = _BORDER
        label_cell.alignment = _ALIGN_CENTER

        desc_cell  = ws.cell(row=i, column=2, value=desc)
        desc_cell.font      = Font(name="Segoe UI", size=10)
        desc_cell.alignment = _ALIGN_WRAP
        desc_cell.border    = _BORDER


# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Feature Activity Tracker v3",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
/* ── Hide Streamlit default UI chrome ── */
#MainMenu { visibility: hidden; }
footer    { visibility: hidden; }
header    { visibility: hidden; }
[data-testid="stDeployButton"]        { display: none !important; }
[data-testid="stStatusWidget"]        { display: none !important; }
[data-testid="manage-app-button"]     { display: none !important; }

/* ── Global ── */
html, body, [class*="css"] { font-family: 'Segoe UI', sans-serif; }

/* ── Project name header ── */
.project-header {
    font-size: 1.9rem;
    font-weight: 700;
    color: #ffffff;
    padding: 2px 0 6px 0;
    border-bottom: 3px solid #1F3864;
    margin-bottom: 14px;
    line-height: 1.2;
}

/* ── Phase header badges ── */
.phase-badge {
    display: inline-block;
    padding: 4px 16px;
    border-radius: 4px;
    font-weight: 700;
    font-size: 1.1rem;
    letter-spacing: 2px;
    color: #ffffff;
    margin-bottom: 6px;
}

/* ── Activity card ── */
.activity-card {
    background: #f8f9fb;
    border-left: 4px solid #1F3864;
    border-radius: 6px;
    padding: 10px 14px;
    margin-bottom: 6px;
}
.activity-card.complete {
    background: #eaf5ec;
    border-left-color: #1F6B2E;
    opacity: 0.75;
}

/* ── Progress bar label ── */
.prog-label { font-size: 0.82rem; color: #555; margin-bottom: 2px; }

/* ── Thin divider ── */
hr.thin { border: none; border-top: 1px solid #dde; margin: 14px 0; }

/* ── Save status badge ── */
.save-saved   { color: #1F6B2E; font-weight: 600; }
.save-unsaved { color: #b85c00; font-weight: 600; }

/* ── Not-needed row ── */
.nn-text { color: #888 !important; text-decoration: line-through; font-style: italic; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────────
if "projects" not in st.session_state:
    raw = load_data()
    st.session_state.projects        = raw["projects"]
    st.session_state.last_saved      = raw.get("last_saved")
    st.session_state.unsaved_changes = False

if "unsaved_changes" not in st.session_state:
    st.session_state.unsaved_changes = False


# ─────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────
def parse_date(val) -> date | None:
    if not val:
        return None
    if isinstance(val, date):
        return val
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(str(val), fmt).date()
        except ValueError:
            pass
    return None


def fmt_date(val) -> str:
    d = parse_date(val)
    return d.strftime("%m/%d/%Y") if d else "—"


def mark_unsaved() -> None:
    """Mark changes and immediately autosave."""
    st.session_state.unsaved_changes = True
    do_save()


def do_save() -> None:
    ts = save_data(st.session_state.projects)
    st.session_state.last_saved      = ts
    st.session_state.unsaved_changes = False


# ─────────────────────────────────────────────────────────────
# ACTIVITY FORM  (add / edit)
# ─────────────────────────────────────────────────────────────
def activity_form(activity: dict | None, form_key: str, activities: list) -> dict | None | str:
    """
    Renders an add/edit form.
    Returns:
      - updated dict on save
      - "CANCEL" when the user cancels
      - None while the form is untouched
    """
    is_new = activity is None
    a = activity or {
        "id":               next_id(activities),
        "phase":            "DISCOVER",
        "feature_activity": "",
        "purpose":          "",
        "responsible":      "",
        "start_date":       "",
        "due_date":         "",
        "notes":            "",
        "complete":         False,
    }

    with st.form(key=form_key, border=True):
        col1, col2 = st.columns(2)
        with col1:
            phase            = st.selectbox(
                "Phase *",
                options=["DISCOVER", "DESIGN", "DELIVER"],
                index=["DISCOVER", "DESIGN", "DELIVER"].index(a["phase"]),
            )
            feature_activity = st.text_input("Feature Activity *", value=a["feature_activity"])
            purpose          = st.text_input("Purpose",            value=a["purpose"])
            _resp_val        = a.get("responsible", "") or ""
            _resp_idx        = RESPONSIBLE_OPTIONS.index(_resp_val) if _resp_val in RESPONSIBLE_OPTIONS else 0
            responsible      = st.selectbox("Responsible", options=RESPONSIBLE_OPTIONS, index=_resp_idx)
        with col2:
            start_date = st.date_input("Start Date", value=parse_date(a["start_date"]), format="MM/DD/YYYY")
            due_date   = st.date_input("Due Date",   value=parse_date(a["due_date"]),   format="MM/DD/YYYY")
            complete   = st.checkbox("✅ Mark as Complete", value=a["complete"])
            notes      = st.text_area("Notes", value=a.get("notes", ""), height=100)

        c1, c2 = st.columns([1, 5])
        submitted = c1.form_submit_button("Add Activity" if is_new else "Save Changes", type="primary")
        cancelled = c2.form_submit_button("Cancel")

        if submitted:
            if not feature_activity.strip():
                st.warning("Feature Activity is required.")
                return None
            return {
                **a,
                "phase":            phase,
                "feature_activity": feature_activity.strip(),
                "purpose":          purpose.strip(),
                "responsible":      responsible.strip(),
                "start_date":       str(start_date) if start_date else "",
                "due_date":         str(due_date)   if due_date   else "",
                "notes":            notes.strip(),
                "complete":         complete,
            }
        if cancelled:
            return "CANCEL"
    return None


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("📋 Activity Tracker")
    st.caption("Feature Activity Checklist  ·  v3")
    st.divider()

    # ── Save ─────────────────────────────────────────────────
    st.subheader("💾 Save")

    if st.button("💾 Save All Projects", type="primary", use_container_width=True):
        do_save()
        st.success("All projects saved!")

    if st.session_state.unsaved_changes:
        st.markdown("<span class='save-unsaved'>🟡 Unsaved changes</span>", unsafe_allow_html=True)
    else:
        st.markdown("<span class='save-saved'>🟢 All saved</span>", unsafe_allow_html=True)

    if st.session_state.last_saved:
        st.caption(f"Last saved: {st.session_state.last_saved}")

    st.divider()

    # ── Add new project tab ──────────────────────────────────
    st.subheader("📁 Project Tabs")
    new_proj_name = st.text_input(
        "New project name",
        placeholder="e.g. Project Alpha",
        key="new_proj_name_input",
    )
    if st.button("➕ Add New Project Tab", use_container_width=True, type="primary"):
        name = new_proj_name.strip() or f"Project {len(st.session_state.projects) + 1}"
        st.session_state.projects.append(make_default_project(name))
        mark_unsaved()
        st.rerun()

    st.divider()

    # ── Overall progress ─────────────────────────────────────
    st.subheader("📊 Overall Progress")
    all_acts  = [a for p in st.session_state.projects for a in p["activities"] if not a.get("not_needed", False)]
    total_all = len(all_acts)
    done_all  = sum(1 for a in all_acts if a["complete"])
    pct_all   = int(done_all / total_all * 100) if total_all else 0
    st.progress(pct_all / 100)
    st.caption(f"{done_all} / {total_all} complete  ({pct_all}%)")

    st.divider()

    # ── Export all projects to Excel ─────────────────────────
    xlsx_all = build_excel_bytes_all(st.session_state.projects)
    st.download_button(
        label="📥 Export All to Excel",
        data=xlsx_all,
        file_name=f"activities_all_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ─────────────────────────────────────────────────────────────
# MAIN AREA
# ─────────────────────────────────────────────────────────────
st.markdown("## 📋 Feature Activity Checklist Tracker")
st.caption(f"Today: {date.today().strftime('%B %d, %Y')}")

projects = st.session_state.projects

if not projects:
    st.info("No projects yet. Add one in the sidebar ➡️")
    st.stop()

# ── Project Tabs ─────────────────────────────────────────────
tab_labels = [p["name"] for p in projects]
tabs       = st.tabs(tab_labels)

PHASE_ORDER = ["DISCOVER", "DESIGN", "DELIVER"]

for tab_idx, (tab, proj) in enumerate(zip(tabs, projects)):
    pid = proj["id"]

    sk_edit   = f"edit_id_{pid}"
    sk_add    = f"add_mode_{pid}"
    sk_rename = f"rename_mode_{pid}"

    for k, default in [
        (sk_edit,   None),
        (sk_add,    False),
        (sk_rename, False),
    ]:
        if k not in st.session_state:
            st.session_state[k] = default

    acts = proj["activities"]

    with tab:
        # ── Project name header ──────────────────────────────
        col_title, col_rename_btn, col_del_btn = st.columns([5, 1, 1])

        with col_title:
            if st.session_state[sk_rename]:
                new_name = st.text_input(
                    "Project name",
                    value=proj["name"],
                    key=f"proj_name_input_{pid}",
                    label_visibility="collapsed",
                )
            else:
                st.markdown(
                    f"<div class='project-header'>📁 {proj['name']}</div>",
                    unsafe_allow_html=True,
                )

        with col_rename_btn:
            if st.session_state[sk_rename]:
                if st.button("✅ Save Name", key=f"save_name_{pid}", use_container_width=True):
                    proj["name"] = new_name.strip() or proj["name"]
                    st.session_state[sk_rename] = False
                    mark_unsaved()
                    st.rerun()
            else:
                if st.button("✏️ Rename", key=f"rename_{pid}", use_container_width=True):
                    st.session_state[sk_rename] = True
                    st.rerun()

        with col_del_btn:
            if len(projects) > 1:
                if st.button("🗑️ Delete Tab", key=f"del_proj_{pid}", use_container_width=True):
                    st.session_state.projects = [
                        p for p in st.session_state.projects if p["id"] != pid
                    ]
                    mark_unsaved()
                    st.rerun()

        st.caption(f"Today: {date.today().strftime('%B %d, %Y')}")
        st.markdown("<hr class='thin'>", unsafe_allow_html=True)

        # ── Per-project progress bar ─────────────────────────
        active_acts = [a for a in acts if not a.get("not_needed", False)]
        total = len(active_acts)
        done  = sum(1 for a in active_acts if a["complete"])
        pct   = int(done / total * 100) if total else 0
        st.progress(pct / 100)
        st.caption(f"{done} / {total} complete  ({pct}%)")

        # ── Add activity ─────────────────────────────────────
        search_term     = ""
        all_phases      = sorted({a["phase"] for a in acts})
        selected_phases = all_phases
        status_filter   = "All"

        btn_c1, btn_c2 = st.columns([2, 7])
        with btn_c1:
            if st.button("➕ Add Activity", key=f"add_btn_{pid}", type="primary", use_container_width=True):
                st.session_state[sk_add]  = True
                st.session_state[sk_edit] = None

        if st.session_state[sk_add]:
            st.subheader("➕ New Activity")
            result = activity_form(None, f"add_form_{pid}", acts)
            if result == "CANCEL":
                st.session_state[sk_add] = False
                st.rerun()
            elif result is not None:
                acts.append(result)
                mark_unsaved()
                st.session_state[sk_add] = False
                st.success(f"Activity **{result['feature_activity']}** added!")
                st.rerun()
            st.markdown("---")

        def matches(a: dict, _phases=selected_phases, _status=status_filter, _search=search_term) -> bool:
            if a["phase"] not in _phases:
                return False
            if _status == "Complete"   and not a["complete"]:
                return False
            if _status == "Incomplete" and a["complete"]:
                return False
            if _search:
                needle   = _search.lower()
                haystack = " ".join([
                    a.get("feature_activity", ""),
                    a.get("purpose",          ""),
                    a.get("responsible",      ""),
                    a.get("notes",            ""),
                ]).lower()
                if needle not in haystack:
                    return False
            return True

        filtered = [a for a in acts if matches(a)]

        # ── Render activities grouped by phase ───────────────
        for phase in PHASE_ORDER:
            phase_acts = [a for a in filtered if a["phase"] == phase]
            if not phase_acts:
                continue

            color       = PHASE_COLORS.get(phase, "#333")
            icon        = PHASE_ICONS.get(phase, "")
            active_phase = [a for a in phase_acts if not a.get("not_needed", False)]
            done_count   = sum(1 for a in active_phase if a["complete"])
            total_count  = len(active_phase)

            st.markdown(
                f"<div class='phase-badge' style='background:{color}'>"
                f"{icon} {phase} &nbsp;·&nbsp; {done_count}/{total_count} complete"
                f"</div>",
                unsafe_allow_html=True,
            )

            hcols = st.columns([0.4, 2.0, 2.6, 1.2, 1.5, 1.5, 0.4, 0.8, 0.8])
            for col, label in zip(
                hcols,
                ["Done", "Feature Activity", "Purpose", "Responsible", "Start", "Due", "", "", ""]
            ):
                col.markdown(
                    f"<span style='font-size:1.05rem;font-weight:700;'>{label}</span>",
                    unsafe_allow_html=True,
                )
            st.markdown("<hr class='thin'>", unsafe_allow_html=True)

            for a in phase_acts:
                if st.session_state[sk_edit] == a["id"]:
                    result = activity_form(a, f"edit_form_{pid}_{a['id']}", acts)
                    if result == "CANCEL":
                        st.session_state[sk_edit] = None
                        st.rerun()
                    elif result is not None:
                        idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                        acts[idx] = result
                        mark_unsaved()
                        st.session_state[sk_edit] = None
                        st.success("Activity saved!")
                        st.rerun()
                    continue

                cols = st.columns([0.4, 2.0, 2.6, 1.2, 1.5, 1.5, 0.4, 0.8, 0.8])
                is_not_needed = a.get("not_needed", False)

                if is_not_needed:
                    cols[0].markdown("🚫", help="Marked as not needed")
                else:
                    new_complete = cols[0].checkbox(
                        "complete",
                        value=a["complete"],
                        key=f"chk_{pid}_{a['id']}",
                        label_visibility="collapsed",
                    )
                    if new_complete != a["complete"]:
                        idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                        acts[idx]["complete"] = new_complete
                        mark_unsaved()
                        st.rerun()

                if is_not_needed:
                    nn_style = "color:#888;text-decoration:line-through;font-style:italic;"
                    cols[1].markdown(f"<span style='{nn_style}'>{a['feature_activity']}</span>", unsafe_allow_html=True)
                    cols[2].markdown(f"<span style='{nn_style}'>{a.get('purpose','')}</span>", unsafe_allow_html=True)
                    cols[3].markdown(f"<span style='color:#888;'>N/A</span>", unsafe_allow_html=True)
                    cols[4].markdown(f"<span style='color:#888;'>N/A</span>", unsafe_allow_html=True)
                    cols[5].markdown(f"<span style='color:#888;'>N/A</span>", unsafe_allow_html=True)
                else:
                    stored_start = parse_date(a.get("start_date"))
                    stored_due   = parse_date(a.get("due_date"))
                    overdue = (
                        stored_due is not None
                        and stored_due < date.today()
                        and not a["complete"]
                    )
                    fa_text  = f"~~{a['feature_activity']}~~" if a["complete"] else a["feature_activity"]
                    pur_text = f"~~{a.get('purpose', '')}~~"  if a["complete"] else a.get("purpose", "")
                    if overdue:
                        cols[1].markdown(
                            f"<span style='color:#ff4b4b;font-weight:600;'>⚠️ {a['feature_activity']}</span>",
                            unsafe_allow_html=True,
                        )
                    else:
                        cols[1].markdown(fa_text)
                    cols[2].markdown(pur_text)

                    # ── Inline Responsible picklist ────────────────────────
                    _cur_resp  = a.get("responsible", "") or ""
                    _resp_idx  = RESPONSIBLE_OPTIONS.index(_cur_resp) if _cur_resp in RESPONSIBLE_OPTIONS else 0
                    new_resp   = cols[3].selectbox(
                        "Responsible",
                        options=RESPONSIBLE_OPTIONS,
                        index=_resp_idx,
                        key=f"resp_{pid}_{a['id']}",
                        label_visibility="collapsed",
                    )
                    if new_resp != _cur_resp:
                        idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                        acts[idx]["responsible"] = new_resp
                        mark_unsaved()
                        st.rerun()

                    # ── Inline Start date picker ──────────────────────────
                    new_start = cols[4].date_input(
                        "Start",
                        value=stored_start,
                        key=f"start_{pid}_{a['id']}",
                        format="MM/DD/YYYY",
                        label_visibility="collapsed",
                    )
                    if new_start != stored_start:
                        idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                        acts[idx]["start_date"] = str(new_start) if new_start else ""
                        mark_unsaved()
                        st.rerun()

                    # ── Inline Due date picker (red label when overdue) ───
                    new_due = cols[5].date_input(
                        "⚠️ OVERDUE" if overdue else "Due",
                        value=stored_due,
                        key=f"due_{pid}_{a['id']}",
                        format="MM/DD/YYYY",
                        label_visibility="visible" if overdue else "collapsed",
                    )
                    if new_due != stored_due:
                        idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                        acts[idx]["due_date"] = str(new_due) if new_due else ""
                        mark_unsaved()
                        st.rerun()

                has_notes = bool(a.get("notes", "").strip())
                cols[6].markdown("📝" if has_notes else "")

                if cols[7].button("✏️", key=f"edit_{pid}_{a['id']}", help="Edit this activity"):
                    st.session_state[sk_edit] = a["id"]
                    st.session_state[sk_add]  = False
                    st.rerun()

                is_nn    = a.get("not_needed", False)
                nn_label = "↩️" if is_nn else "🚫"
                nn_help  = "Mark as needed again" if is_nn else "Mark as not needed"
                if cols[8].button(nn_label, key=f"nn_{pid}_{a['id']}", help=nn_help):
                    idx = next(i for i, x in enumerate(acts) if x["id"] == a["id"])
                    acts[idx]["not_needed"] = not is_nn
                    if not is_nn:
                        acts[idx]["complete"] = False
                    mark_unsaved()
                    st.rerun()

                if has_notes:
                    with st.expander("📝 Notes", expanded=False):
                        st.info(a["notes"])

            st.markdown("<br>", unsafe_allow_html=True)

        if not filtered:
            st.info("No activities match your current filters.")

        # ── Per-tab Excel export ─────────────────────────────
        st.markdown("---")
        safe_name  = proj["name"].replace(" ", "_")
        xlsx_tab   = build_excel_bytes(proj["activities"], sheet_name=proj["name"])
        st.download_button(
            label=f"📥 Export '{proj['name']}' to Excel",
            data=xlsx_tab,
            file_name=f"{safe_name}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"export_{pid}",
        )
